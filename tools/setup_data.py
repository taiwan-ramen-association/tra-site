"""
setup_data.py — 資料處理工具
A  開始編輯：JSON → Excel → 開啟檔案
B  完成編輯：Excel → JSON → 正規化 → 寫回 Excel
0  進階單步執行
"""
import csv
import json
import os
import re
import subprocess
import sys
import time
import warnings
import xml.etree.ElementTree as ET

warnings.filterwarnings('ignore')

def install(pkg):
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg, '-q'])

try:
    import requests
except ImportError:
    print('安裝 requests 中...')
    install('requests')
    import requests

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, numbers
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    print('安裝 openpyxl 中...')
    install('openpyxl')
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, numbers
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

tools_dir = os.path.dirname(os.path.abspath(__file__))
root_dir  = os.path.dirname(tools_dir)
json_path = os.path.join(root_dir, 'data', 'data.json')
xlsx_path = os.path.join(tools_dir, 'data.xlsx')
dist_path = os.path.join(root_dir, 'data', 'districts.json')

# ── 共用 I/O ──────────────────────────────────────────────────────────────────
def load_data():
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_data(rows):
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)

def load_districts():
    if not os.path.exists(dist_path):
        return {}
    with open(dist_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def section(num, title):
    print()
    print('─' * 54)
    print(f'  [{num}]  {title}')
    print('─' * 54)

# ════════════════════════════════════════════════════════════════════════════════
# STEP 1：更新行政區清單
# ════════════════════════════════════════════════════════════════════════════════
def step_update_districts():
    section(1, '更新行政區清單（內政部 API）')

    API_BASE = 'https://api.nlsc.gov.tw/other/ListTown1'
    districts_raw = {}

    for code in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
        try:
            r = requests.get(f'{API_BASE}/{code}', timeout=10, verify=False)
            if r.status_code != 200 or '<townItem>' not in r.text:
                continue
            root  = ET.fromstring(r.content)
            items = root.findall('townItem')
            if items:
                districts_raw[code] = [item.findtext('townname') for item in items]
                print(f'    {code}: {len(districts_raw[code])} 個鄉鎮市區')
        except Exception as e:
            print(f'    {code}: 失敗 ({e})')

    county_names = {}
    try:
        r = requests.get('https://api.nlsc.gov.tw/other/ListCounty', timeout=10, verify=False)
        root = ET.fromstring(r.content)
        for item in root.findall('countyItem'):
            county_names[item.findtext('countycode')] = item.findtext('countyname')
    except Exception as e:
        print(f'    縣市名稱 API 失敗: {e}')

    districts = {county_names.get(c, c): towns for c, towns in districts_raw.items()}
    with open(dist_path, 'w', encoding='utf-8') as f:
        json.dump(districts, f, ensure_ascii=False, indent=2)

    total = sum(len(v) for v in districts.values())
    print(f'\n  ✅ 完成：{len(districts)} 縣市，{total} 鄉鎮市區')
    return len(districts)

# ════════════════════════════════════════════════════════════════════════════════
# STEP 2：補縣市／鄉鎮市區
# ════════════════════════════════════════════════════════════════════════════════
def step_fill_city_district():
    section(2, '補縣市／鄉鎮市區')

    districts = load_districts()
    if not districts:
        print('  ⚠  找不到 districts.json，請先執行步驟 1')
        return 0

    rows = load_data()

    def parse(addr):
        if not addr:
            return '', ''
        s = re.sub(r'^\d{3,6}', '', addr.replace('台', '臺')).strip()
        for county, towns in districts.items():
            if s.startswith(county):
                rest = s[len(county):]
                for town in towns:
                    if rest.startswith(town):
                        return county, town
        return '', ''

    updated = 0
    for row in rows:
        addr = row.get('地址', '')
        if addr:
            cleaned = re.sub(r'^\d{3,6}', '', addr).strip()
            if cleaned != addr:
                row['地址'] = cleaned

        # 永遠重新從地址解析，若有新值則覆蓋（處理行政區升格等名稱異動）
        city, dist = parse(row.get('地址', ''))
        changed = False
        if city and row.get('縣市') != city:
            print(f'    ✓ {row["店名"]}：縣市 {row.get("縣市","（空）")!r} → {city!r}')
            row['縣市'] = city
            changed = True
        if dist and row.get('鄉鎮市區') != dist:
            print(f'    ✓ {row["店名"]}：鄉鎮市區 {row.get("鄉鎮市區","（空）")!r} → {dist!r}')
            row['鄉鎮市區'] = dist
            changed = True
        if changed:
            updated += 1
        elif not city:
            print(f'    ✗ {row["店名"]}：地址解析失敗')

    save_data(rows)
    print(f'\n  ✅ 完成：補填 {updated} 筆（共 {len(rows)} 筆）')
    return updated

# ════════════════════════════════════════════════════════════════════════════════
# STEP 3：補 lat/lng 座標
# ════════════════════════════════════════════════════════════════════════════════
def step_geocode():
    section(3, '補 lat/lng 座標')

    print('  模式選擇（直接 Enter = 只補缺少座標）：')
    print('    1. 只補缺少座標的店家  ← 預設')
    print('    2. 重新更正所有有 Map URL 的店家（修正舊座標精度）')
    mode = input('  請輸入 1 或 2：').strip() or '1'

    rows  = load_data()
    total = len(rows)
    if mode == '2':
        to_geocode = [r for r in rows if r.get('Map', '').startswith('http') or not r.get('lat')]
        print(f'  重新 geocode：{len(to_geocode)} 筆（共 {total} 筆）')
    else:
        to_geocode = [r for r in rows if not r.get('lat') or not r.get('lng')]
        print(f'  需要 geocode：{len(to_geocode)} 筆（共 {total} 筆）')

    if not to_geocode:
        print('  ✅ 無需處理')
        return 0

    UA = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
    updated    = 0
    failed     = []
    consecutive = 0
    MAX_CONSEC  = 5

    def from_map_url(url):
        if not url or not url.startswith('http'):
            return None, None
        r = requests.get(url, headers=UA, timeout=10, verify=False, allow_redirects=True)
        # !3d!4d 是 Google Maps 標記點的精確座標
        # 不使用 /@ 的 fallback：那是地圖視角中心，縮放狀態不同會漂移到海上
        m = re.search(r'!3d(-?\d+\.\d+)!4d(-?\d+\.\d+)', r.url)
        if m:
            return float(m.group(1)), float(m.group(2))
        return None, None

    def from_nominatim(address):
        r = requests.get('https://nominatim.openstreetmap.org/search',
            params={'q': address, 'format': 'json', 'limit': 1},
            headers=UA, timeout=10, verify=False)
        res = r.json()
        return (float(res[0]['lat']), float(res[0]['lon'])) if res else (None, None)

    for i, row in enumerate(to_geocode):
        name    = row.get('店名', '')
        address = row.get('地址', '') or name
        print(f'  [{i+1}/{len(to_geocode)}] {name}')
        try:
            lat, lng = from_map_url(row.get('Map', ''))
            if lat:
                print(f'    ✓ (Map URL) {lat:.6f}, {lng:.6f}')
            else:
                lat, lng = from_nominatim(address)
                if lat:
                    print(f'    ✓ (Nominatim) {lat:.6f}, {lng:.6f}')

            if lat:
                row['lat'] = lat
                row['lng'] = lng
                updated   += 1
                consecutive = 0
            else:
                failed.append(name)
                consecutive += 1
                print(f'    ✗ 找不到座標（連續失敗 {consecutive}/{MAX_CONSEC}）')
        except Exception as e:
            failed.append(name)
            consecutive += 1
            print(f'    ✗ 錯誤：{e}')

        if consecutive >= MAX_CONSEC:
            print(f'\n  ⚠  連續失敗 {MAX_CONSEC} 筆，中斷作業')
            break
        time.sleep(1.1)

    save_data(rows)
    print(f'\n  ✅ 完成：更新 {updated} 筆（共 {total} 筆）')
    if failed:
        print(f'  ⚠  無法取得座標：{", ".join(failed)}')
    return updated

# ════════════════════════════════════════════════════════════════════════════════
# STEP 4：正規化營業時段
# ════════════════════════════════════════════════════════════════════════════════
HOURS_FIELDS = ['週一', '週二', '週三', '週四', '週五', '週六', '週日', '營業時段']

def normalize_hours(value):
    if not isinstance(value, str) or not value.strip():
        return value
    v = value.strip()
    v = re.sub(r'(?<=\d)[—\-~～](?=\d)', '–', v)
    segments = re.findall(r'\d{1,2}:\d{2}–\d{1,2}:\d{2}', v)
    return '、'.join(segments) if segments else v

def step_normalize_hours():
    section(4, '正規化營業時段格式')

    rows    = load_data()
    updated = 0
    for row in rows:
        for field in HOURS_FIELDS:
            original   = row.get(field, '')
            normalized = normalize_hours(original)
            if normalized != original:
                row[field] = normalized
                updated   += 1
                print(f'    {row["店名"]} [{field}]  {original!r} → {normalized!r}')

    save_data(rows)
    print(f'\n  ✅ 完成：更新 {updated} 個欄位（共 {len(rows)} 筆）')
    return updated

# ════════════════════════════════════════════════════════════════════════════════
# STEP 5：正規化星期排序
# ════════════════════════════════════════════════════════════════════════════════
DAY_ORDER  = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '日': 7}
DAY_FIELDS = ['營業日', '店休日']

def normalize_days(value):
    if not value or not isinstance(value, str):
        return value
    parts = [p.strip() for p in value.split(',')]
    if not all(p in DAY_ORDER for p in parts if p):
        return value
    return ', '.join(sorted(parts, key=lambda d: DAY_ORDER.get(d, 99)))

def step_normalize_days():
    section(5, '正規化星期排序')

    rows    = load_data()
    updated = 0
    for row in rows:
        for field in DAY_FIELDS:
            original   = row.get(field, '')
            normalized = normalize_days(original)
            if normalized != original:
                row[field] = normalized
                updated   += 1
                print(f'    {row["店名"]} [{field}]  {original!r} → {normalized!r}')

    save_data(rows)
    print(f'\n  ✅ 完成：更新 {updated} 個欄位（共 {len(rows)} 筆）')
    return updated

# ════════════════════════════════════════════════════════════════════════════════
# STEP 6：正規化開幕日期
# ════════════════════════════════════════════════════════════════════════════════
DATE_FIELDS = ['開幕日', '歇業日']

def normalize_date(value):
    """各種日期格式統一為 YYYY-MM-DD，無法辨識則原樣返回"""
    if not value or not isinstance(value, str) or not value.strip():
        return value
    v = value.strip()

    # 已經是標準格式
    if re.match(r'^\d{4}-\d{2}-\d{2}$', v):
        return v

    # YYYY-MM-DD HH:MM:SS 或 YYYY-MM-DD HH:MM（Excel datetime）
    m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})[\sT]\d{1,2}:\d{2}', v)
    if m:
        return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'

    # YYYY/MM/DD 或 YYYY/M/D
    m = re.match(r'^(\d{4})/(\d{1,2})/(\d{1,2})$', v)
    if m:
        return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'

    # YYYY.MM.DD 或 YYYY.M.D
    m = re.match(r'^(\d{4})\.(\d{1,2})\.(\d{1,2})$', v)
    if m:
        return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'

    # YYYY-M-D（有破折號但未補零）
    m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', v)
    if m:
        return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'

    # YYYYMMDD（純數字 8 碼）
    m = re.match(r'^(\d{4})(\d{2})(\d{2})$', v)
    if m:
        return f'{m.group(1)}-{m.group(2)}-{m.group(3)}'

    return v  # 無法辨識，原樣返回

def step_normalize_dates():
    section(6, '正規化開幕日 / 歇業日（→ YYYY-MM-DD）')

    rows    = load_data()
    updated = 0
    failed  = []

    for row in rows:
        for field in DATE_FIELDS:
            original = row.get(field, '')
            if not original:
                continue
            normalized = normalize_date(original)
            if normalized == original:
                continue
            if re.match(r'^\d{4}-\d{2}-\d{2}$', normalized):
                row[field] = normalized
                updated   += 1
                print(f'    ✓ {row["店名"]} [{field}]  {original!r} → {normalized!r}')
            else:
                failed.append((row['店名'], field, original))

        # 開幕日正規化後同步更新開幕月份
        d = str(row.get('開幕日', '')).strip()
        if re.match(r'^\d{4}-\d{2}-\d{2}$', d):
            month = int(d.split('-')[1])
            if row.get('開幕月份') != month:
                print(f'    ✓ {row["店名"]}：開幕月份 {row.get("開幕月份","（空）")} → {month}')
                row['開幕月份'] = month
                updated += 1

    save_data(rows)
    print(f'\n  ✅ 完成：更新 {updated} 個欄位（共 {len(rows)} 筆）')
    if failed:
        print(f'  ⚠  無法辨識格式（請手動修正）：')
        for name, field, val in failed:
            print(f'      {name} [{field}] = {val!r}')
    return updated

# ════════════════════════════════════════════════════════════════════════════════
# STEP 7：分配店家 ID
# ════════════════════════════════════════════════════════════════════════════════
CITY_CODE = {
    '臺北市': 'A', '台北市': 'A',
    '新北市': 'B',
    '桃園市': 'C',
    '臺中市': 'D', '台中市': 'D',
    '臺南市': 'E', '台南市': 'E',
    '高雄市': 'F',
    '基隆市': 'G',
    '新竹市': 'H',
    '新竹縣': 'I',
    '苗栗縣': 'J',
    '彰化縣': 'K',
    '南投縣': 'L',
    '雲林縣': 'M',
    '嘉義市': 'N',
    '嘉義縣': 'O',
    '屏東縣': 'P',
    '宜蘭縣': 'Q',
    '花蓮縣': 'R',
    '臺東縣': 'S', '台東縣': 'S',
    '澎湖縣': 'T',
    '金門縣': 'U',
    '連江縣': 'V',
}
ID_RE = re.compile(r'^[A-Z]\d{5}$')

CODE_TO_CITY = {
    'A': '臺北市', 'B': '新北市', 'C': '桃園市', 'D': '臺中市',
    'E': '臺南市', 'F': '高雄市', 'G': '基隆市', 'H': '新竹市',
    'I': '新竹縣', 'J': '苗栗縣', 'K': '彰化縣', 'L': '南投縣',
    'M': '雲林縣', 'N': '嘉義市', 'O': '嘉義縣', 'P': '屏東縣',
    'Q': '宜蘭縣', 'R': '花蓮縣', 'S': '臺東縣', 'T': '澎湖縣',
    'U': '金門縣', 'V': '連江縣', 'Z': '未知縣市',
}

def _get_city_for_id(row):
    city = str(row.get('縣市', '')).strip()
    if not city:
        addr = str(row.get('地址', '')).strip()
        addr = re.sub(r'^\d{3,6}', '', addr)
        city = addr[:3]
    return city.replace('台', '臺')

counters_path = os.path.join(root_dir, 'data', 'id_counters.json')

def _load_counters():
    if os.path.exists(counters_path):
        with open(counters_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def _save_counters(city_max):
    with open(counters_path, 'w', encoding='utf-8') as f:
        json.dump(city_max, f, ensure_ascii=False, indent=2, sort_keys=True)

def step_assign_ids():
    section(7, '分配店家 ID（縣市代碼 + 5位流水號）')

    rows = load_data()

    # 從 id_counters.json 載入歷史最大值（防止因資料列被誤刪導致 ID 重用）
    city_max = _load_counters()
    print(f'  📋 讀取歷史計數器：{dict(sorted(city_max.items()))}')

    # 再與現有資料的 ID 取最大值（兩者都納入，只增不減）
    for row in rows:
        eid = str(row.get('ID', '')).strip()
        if ID_RE.match(eid):
            letter = eid[0]
            city_max[letter] = max(city_max.get(letter, 0), int(eid[1:]))

    assigned = 0
    for row in rows:
        eid = str(row.get('ID', '')).strip()
        if ID_RE.match(eid):
            continue
        city   = _get_city_for_id(row)
        letter = CITY_CODE.get(city, 'Z')
        next_n = city_max.get(letter, 0) + 1
        city_max[letter] = next_n
        row['ID'] = f'{letter}{next_n:05d}'
        assigned += 1
        print(f'    ✓ {row["店名"]}  →  {row["ID"]}')

    # 確保 ID 排在第一欄
    rows = [{'ID': r.get('ID', ''), **{k: v for k, v in r.items() if k != 'ID'}} for r in rows]

    save_data(rows)

    # 寫回計數器（只增不減，是唯一的安全防線）
    _save_counters(city_max)
    print(f'  💾 計數器已更新 → data/id_counters.json')

    print(f'\n  ✅ 完成：新分配 {assigned} 筆，共 {len(rows)} 筆')

    city_counts = {}
    for row in rows:
        letter = row.get('ID', 'Z')[0] if row.get('ID') else 'Z'
        city_counts[letter] = city_counts.get(letter, 0) + 1
    print('\n  各縣市店家數量：')
    for letter in sorted(city_counts):
        print(f'    {letter} {CODE_TO_CITY.get(letter, letter)}: {city_counts[letter]} 間')

    return assigned

# ════════════════════════════════════════════════════════════════════════════════
# Excel ↔ JSON 轉換
# ════════════════════════════════════════════════════════════════════════════════
DATE_TEXT_FIELDS = {'開幕日', 'ID'}

def step_excel_to_json():
    section('E', 'Excel → JSON')
    if not os.path.exists(xlsx_path):
        print('  ❌ 找不到 data.xlsx，請先執行 A【開始編輯】')
        return False
    print('  📂 讀取 data.xlsx...')
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    rows_raw = list(ws.values)
    if not rows_raw:
        print('  ❌ Excel 是空的')
        return False
    headers = [str(h).strip() for h in rows_raw[0]]
    rows = []
    for row in rows_raw[1:]:
        if all((v is None or str(v).strip() == '') for v in row):
            continue
        obj = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else ''
            if val is None:
                val = ''
            else:
                val = str(val).strip()
                if val.endswith('.0') and val[:-2].lstrip('-').isdigit():
                    val = val[:-2]
            obj[h] = val
        if obj.get('店名', '').strip():
            rows.append(obj)
    save_data(rows)
    print(f'  ✅ 完成：data.json 已更新（共 {len(rows)} 筆）')
    return True

def step_json_to_excel():
    section('X', 'JSON → Excel（含樣式與下拉驗證）')
    rows = load_data()
    if not rows:
        print('  ❌ data.json 是空的')
        return False
    print(f'  📝 寫入 {len(rows)} 筆資料...')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '店家資料'
    headers = list(rows[0].keys())
    header_fill = PatternFill(start_color='C8272D', end_color='C8272D', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row_idx, row in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(h, ''))
            if h in DATE_TEXT_FIELDS:
                cell.number_format = numbers.FORMAT_TEXT
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    csv_path = os.path.join(tools_dir, 'item_detail.csv')
    if os.path.exists(csv_path):
        with open(csv_path, 'r', encoding='utf-8-sig') as f:
            reader      = csv.reader(f)
            vld_headers = next(reader)
            vld_cols    = {h: [] for h in vld_headers}
            for r in reader:
                for i, h in enumerate(vld_headers):
                    val = r[i].strip() if i < len(r) else ''
                    if val:
                        vld_cols[h].append(val)
        ws_vld = wb.create_sheet('驗證清單')
        ws_vld.sheet_state = 'hidden'
        for col_idx, h in enumerate(vld_headers, 1):
            ws_vld.cell(row=1, column=col_idx, value=h)
            for row_idx, val in enumerate(vld_cols[h], 2):
                ws_vld.cell(row=row_idx, column=col_idx, value=val)
        last_row   = len(rows) + 1
        col_letter = {}
        for col_idx, h in enumerate(headers, 1):
            if h in vld_cols:
                col_letter[h] = get_column_letter(col_idx)
        for vld_col_idx, h in enumerate(vld_headers, 1):
            letter = col_letter.get(h)
            if not letter or not vld_cols[h]:
                continue
            vld_col_letter = get_column_letter(vld_col_idx)
            vld_range = f'驗證清單!${vld_col_letter}$2:${vld_col_letter}${len(vld_cols[h]) + 1}'
            dv = DataValidation(type='list', formula1=vld_range, showDropDown=False, allow_blank=True)
            dv.sqref = f'{letter}2:{letter}{last_row}'
            ws.add_data_validation(dv)
        print('  ✅ 已套用下拉選單驗證')
    wb.save(xlsx_path)
    print(f'  ✅ 完成！data.xlsx 已產生（共 {len(rows)} 筆）')
    print(f'  📍 {xlsx_path}')
    return True

# ════════════════════════════════════════════════════════════════════════════════
# 選單
# ════════════════════════════════════════════════════════════════════════════════
STEPS = [
    (1, '更新行政區清單（內政部 API）',   step_update_districts),
    (2, '補縣市／鄉鎮市區',               step_fill_city_district),
    (3, '補 lat/lng 座標',                step_geocode),
    (4, '正規化營業時段格式',             step_normalize_hours),
    (5, '正規化星期排序',                 step_normalize_days),
    (6, '正規化開幕日 / 歇業日（→ YYYY-MM-DD）', step_normalize_dates),
    (7, '分配店家 ID',                    step_assign_ids),
]

def show_menu():
    print()
    print('╔' + '═' * 52 + '╗')
    print('║{:^52}║'.format('資料處理工具　Setup Data'))
    print('╠' + '═' * 52 + '╣')
    print('║  A  【拉取最新】git pull{:<28}║'.format(''))
    print('║  B  【開始編輯】JSON → Excel，開啟檔案{:<12}║'.format(''))
    print('║  C  【完成編輯】Excel → JSON → 正規化 → Excel{:<6}║'.format(''))
    print('║  D  【推上遠端】git push data.json + 計數器{:<8}║'.format(''))
    print('║  ' + '─' * 49 + '║')
    print('║  0  進階單步執行{:<35}║'.format(''))
    print('║  ' + '─' * 49 + '║')
    print('║  q  離開{:<43}║'.format(''))
    print('╚' + '═' * 52 + '╝')

def show_advanced_menu():
    print()
    print('╔' + '═' * 52 + '╗')
    print('║{:^52}║'.format('進階單步執行'))
    print('╠' + '═' * 52 + '╣')
    for num, desc, _ in STEPS:
        print(f'║  {num}  {desc:<46}║')
    print('║  ' + '─' * 49 + '║')
    print('║  b  返回主選單{:<37}║'.format(''))
    print('╚' + '═' * 52 + '╝')

def run_path_a():
    print('\n▶ A【拉取最新】git pull')
    result = subprocess.run(
        ['git', 'pull'],
        cwd=root_dir, capture_output=True, text=True, encoding='utf-8'
    )
    print(result.stdout.strip() or result.stderr.strip())
    if result.returncode == 0:
        print('  ✅ 完成')
    else:
        print('  ❌ git pull 失敗')

def run_path_b():
    print('\n▶ B【開始編輯】JSON → Excel → 開啟檔案')
    ok = step_json_to_excel()
    if ok:
        print('\n  📂 開啟 Excel...')
        subprocess.Popen(['cmd', '/c', 'start', '', xlsx_path])

def run_path_d():
    print('\n▶ D【推上遠端】git push data.json + id_counters.json')

    result = subprocess.run(
        ['git', 'add', 'data/data.json', 'data/id_counters.json'],
        cwd=root_dir, capture_output=True, text=True, encoding='utf-8'
    )
    if result.returncode != 0:
        print(f'  ❌ git add 失敗：{result.stderr.strip()}')
        return

    status = subprocess.run(
        ['git', 'diff', '--cached', '--stat'],
        cwd=root_dir, capture_output=True, text=True, encoding='utf-8'
    ).stdout.strip()

    if not status:
        print('  ℹ  無變更，不需要 commit')
        return

    print(f'\n  變更內容：\n{status}\n')
    msg = input('  請輸入 commit 訊息（直接 Enter 使用預設）：').strip()
    if not msg:
        msg = f'更新店家資料'

    result = subprocess.run(
        ['git', 'commit', '-m', msg],
        cwd=root_dir, capture_output=True, text=True, encoding='utf-8'
    )
    if result.returncode != 0:
        print(f'  ❌ commit 失敗：{result.stderr.strip()}')
        return
    print(f'  ✅ Committed')

    print('  🚀 git push...')
    result = subprocess.run(
        ['git', 'push'],
        cwd=root_dir, capture_output=True, text=True, encoding='utf-8'
    )
    if result.returncode == 0:
        print('  ✅ Push 完成！')
    else:
        print(f'  ❌ push 失敗：{result.stderr.strip()}')

def run_path_c():
    print('\n▶ C【完成編輯】Excel → JSON → 正規化 → Excel')
    if not step_excel_to_json():
        return
    step_assign_ids()
    step_fill_city_district()
    step_normalize_hours()
    step_normalize_days()
    step_normalize_dates()
    step_json_to_excel()
    print()
    print('═' * 54)
    print('  ✅ 完成！data.json 與 data.xlsx 均已更新')
    print('═' * 54)

while True:
    show_menu()
    choice = input('\n請輸入選項：').strip().lower()

    if choice == 'q':
        print('\n掰掰')
        break

    elif choice == 'a':
        run_path_a()
        input('\n按 Enter 繼續...')

    elif choice == 'b':
        run_path_b()
        input('\n按 Enter 繼續...')

    elif choice == 'c':
        run_path_c()
        input('\n按 Enter 繼續...')

    elif choice == 'd':
        run_path_d()
        input('\n按 Enter 繼續...')

    elif choice == '0':
        while True:
            show_advanced_menu()
            sub = input('\n請輸入數字（b 返回）：').strip().lower()
            if sub == 'b':
                break
            elif sub.isdigit() and 1 <= int(sub) <= len(STEPS):
                _, _, fn = STEPS[int(sub) - 1]
                fn()
                input('\n按 Enter 繼續...')
            else:
                print(f'\n  ⚠  「{sub}」不是有效的選項')

    else:
        print(f'\n  ⚠  「{choice}」不是有效的選項')
        input('\n按 Enter 繼續...')

"""
json_to_excel.py — 純轉換：JSON → Excel（含樣式與下拉驗證）
正規化請執行 setup_data.py
"""
import csv
import json
import os
import subprocess
import sys

def install(pkg):
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg, '-q'])

try:
    import openpyxl
except ImportError:
    print('安裝 openpyxl 中...')
    install('openpyxl')
    import openpyxl

tools_dir = os.path.dirname(os.path.abspath(__file__))
root_dir  = os.path.dirname(tools_dir)
json_path = os.path.join(root_dir, 'data', 'data.json')
xlsx_path = os.path.join(tools_dir, 'data.xlsx')

print('📂 讀取 data.json...')
with open(json_path, 'r', encoding='utf-8') as f:
    rows = json.load(f)

if not rows:
    print('❌ data.json 是空的')
    input('按 Enter 關閉...')
    sys.exit(1)

# ── 建立 Excel ────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '店家資料'
headers = list(rows[0].keys())

# 標題列樣式
from openpyxl.styles import PatternFill, Font, Alignment, numbers
header_fill = PatternFill(start_color='C8272D', end_color='C8272D', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True, size=11)

# 開幕日欄位強制設為文字格式，防止 Excel 自動轉型
DATE_TEXT_FIELDS = {'開幕日'}

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.fill      = header_fill
    cell.font      = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 資料列
for row_idx, row in enumerate(rows, 2):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=row.get(h, ''))
        if h in DATE_TEXT_FIELDS:
            cell.number_format = numbers.FORMAT_TEXT

# 凍結第一列
ws.freeze_panes = 'A2'

# 自動欄寬（最大 40）
for col in ws.columns:
    max_len = max((len(str(cell.value or '')) for cell in col), default=0)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

# ── 下拉驗證（item_detail.csv）────────────────────────────────────────────────
csv_path = os.path.join(tools_dir, 'item_detail.csv')
if os.path.exists(csv_path):
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader     = csv.reader(f)
        vld_headers = next(reader)
        vld_cols   = {h: [] for h in vld_headers}
        for row in reader:
            for i, h in enumerate(vld_headers):
                val = row[i].strip() if i < len(row) else ''
                if val:
                    vld_cols[h].append(val)

    ws_vld = wb.create_sheet('驗證清單')
    ws_vld.sheet_state = 'hidden'
    for col_idx, h in enumerate(vld_headers, 1):
        ws_vld.cell(row=1, column=col_idx, value=h)
        for row_idx, val in enumerate(vld_cols[h], 2):
            ws_vld.cell(row=row_idx, column=col_idx, value=val)

    last_row   = len(rows) + 1
    col_letter = {}
    for col_idx, h in enumerate(headers, 1):
        if h in vld_cols:
            col_letter[h] = get_column_letter(col_idx)

    for vld_col_idx, h in enumerate(vld_headers, 1):
        letter = col_letter.get(h)
        if not letter or not vld_cols[h]:
            continue
        vld_col_letter = get_column_letter(vld_col_idx)
        vld_range = f'驗證清單!${vld_col_letter}$2:${vld_col_letter}${len(vld_cols[h]) + 1}'
        dv = DataValidation(type='list', formula1=vld_range, showDropDown=False, allow_blank=True)
        dv.sqref = f'{letter}2:{letter}{last_row}'
        ws.add_data_validation(dv)

    print('✅ 已套用下拉選單驗證')

wb.save(xlsx_path)
print(f'✅ 完成！data.xlsx 已產生（共 {len(rows)} 筆）')
print(f'📍 {xlsx_path}')
print()
print('💡 如需正規化資料，請先執行 setup_data.py 再重新轉換')
print()
input('按 Enter 關閉...')
